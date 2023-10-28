import argparse
import os
from typing import Union, Iterable

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from office_tools.pandas_stuff import coerce_df_numtype

pathy = Union[str, os.PathLike]


def edit_excel_batch(infile: pathy, outfile: pathy, sheet: str, header_idx: int, id_header: str, value_data: Iterable,
                     value_header: str,
                     data_insert: str):
    df = pd.read_excel(infile, sheet_name=sheet, header=header_idx)
    coerce_df_numtype(df, col_header=value_header, data=data_insert)

    skipped = []
    for data in value_data:
        if not set_data(df, id_data=data, id_header=id_header, value_header=value_header, value_data=data_insert):
            skipped.append(data)

    if skipped and input(f"Missed: {skipped}\nContinue?") != 'y': exit()

    write_excel(outfile, df, sheet)


def write_excel(outfile: pathy, df: pd.DataFrame, sheet: str):
    try:
        df.to_excel(outfile, sheet_name=sheet, index=False)
    except PermissionError:
        print("Close Excel and retry.")
    except Exception as e:
        print(f"Error: {e}")
    else:
        print("Saved.")
        return 0



def check_data(df: pd.DataFrame, id_data: str | int, id_header: str, value_header: str,
               value_data: str | int | float) -> bool:
    """ Returns True if value_data is in col value_header for row id_data."""
    row = df[df[id_header].astype(str) == str(id_data)]

    if len(row) != 1:
        raise ValueError()
    res = value_data == row[value_header].iloc[0]
    print(f'{value_data} is {"not" if not res else ""} in {value_header} for {id_data}')
    return res


def set_data(df: pd.DataFrame, id_data: str | int, id_header: str, value_header: str, value_data: str | int | float):
    coerce_df_numtype(df=df, col_header=value_header, data=value_data)
    index_to_set = df[df[id_header].astype(str) == str(id_data)].index
    if len(index_to_set) != 1:
        raise ValueError()
    df.at[index_to_set[0], value_header] = value_data
    return True


def get_data_from_excel(df:pd.DataFrame, id_data: str, id_header: str, value_header: str):
    """ Returns data in col value_header for row id_data.
    :param df: DataFrame to search
    :param id_data: Data to search for
    :param id_header: Header for id of record to search
    :param value_header: Column name for data to return
    """
    row = df[df[id_header].str.upper() == id_data.upper()]
    if len(row) != 1:
        print(f"No or multiple results found for {id_data}")
        input("Press enter to continue...")
    actual_data = row[value_header].iloc[0]
    return actual_data


def get_matching(df, key_column, result_column, value):
    result_values = df.loc[df[key_column].astype(str) == value, result_column].values
    if result_values.size == 0 or pd.isna(result_values[0]):
        replacement_value = input(f"No result found for {value}. Enter new value or 's' to skip: ")
        if replacement_value:
            if replacement_value == 's':
                return [None]
            df.loc[df[key_column].astype(str) == value, result_column] = replacement_value
            return replacement_value  # return the new value
        raise ValueError(f"No result found for {value}")
    if result_values.size != 1:
        raise ValueError(f"Multiple results found for {value}: {', '.join(map(str, result_values))}")
    return result_values



def df_overwrite_wb(input_workbook, sheet, df, header_row, out_file):
    wb = load_workbook(input_workbook)
    ws = wb[sheet]
    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + header_row, column=c_idx, value=value)
    wb.save(out_file)


def get_rows(filename, start=0,end=3):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    vals = [i for i in ws.values][start:end]
    return vals



def check_paid(df, id_data):
    return get_data_from_excel(df=df, id_header="No.", value_header="Status", id_data=id_data)



def main(args):
    if os.path.isfile(args.id_data):
        args.id_data = os.path.splitext(os.path.basename(args.id_data))[0]
    df = pd.read_excel(args.workbook_ast, sheet_name='Sales', header=2)
    print(f"Checking {args.workbook_ast} for {args.id_data}")
    result = check_paid(df=df, id_data=args.id_data)
    print(f"Value for {args.id_data} is {result}")
    input("Press enter to exit...")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Check Excel file for data.')
    parser.add_argument('--workbook', help='The Excel file to check')
    parser.add_argument('--id_data', help='List of data to match')
    args = parser.parse_args()
    main(args)
